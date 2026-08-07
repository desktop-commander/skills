from __future__ import annotations

import argparse
import csv
import json
import re
import time
import xml.etree.ElementTree as ET
from pathlib import Path
from urllib.parse import urlparse

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SITEMAP_INDEX_URL = "https://10times.com/xml/sitemaps.xml"
NS = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
REQUEST_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
}
SCHEMA_COLUMNS = [
    "PELNA NAZWA WYDARZENIA", "DATA OD", "DATA DO", "EDYCJE",
    "MIASTO", "KRAJ", "KONTYNENT",
    "ORGANIZATOR NAZWA SKROCONA", "ORGANIZATOR PELNA NAZWA",
    "WYSTAWCY", "WYSTAWCY Z POLSKI", "WYSTAWCY Z KRAJU WYDARZENIA",
    "WWW", "ZAKRES BRANZOWY TARGOW", "PROFIL WYSTAWCY", "PROFIL ODWIEDZAJACEGO",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build 10times public inventory and optional schema export.")
    parser.add_argument("--workspace", required=True, help="Absolute workspace path")
    parser.add_argument("--output-dir", default="", help="Optional absolute output directory")
    parser.add_argument("--local-csv", default="", help="Optional WHR merged CSV path")
    parser.add_argument("--local-raw", default="", help="Optional WHR raw JSON path")
    return parser.parse_args()


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\x00", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return "" if text.lower() in {"", "nan", "none", "null"} else text


def iso_to_pl(value: str) -> str:
    text = clean(value)
    match = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", text)
    return f"{match.group(3)}.{match.group(2)}.{match.group(1)}" if match else ""


def first_nonempty(*values: object) -> str:
    for value in values:
        text = clean(value)
        if text:
            return text
    return ""


def short_organizer_name(name: str) -> str:
    text = clean(name)
    if not text:
        return ""
    for separator in [",", " - ", " | "]:
        if separator in text:
            candidate = clean(text.split(separator)[0])
            if 3 <= len(candidate) <= 60:
                return candidate
    return text if len(text) <= 60 else text[:57].rstrip() + "..."


def unique_join(parts: list[str], sep: str = " | ") -> str:
    seen: set[str] = set()
    ordered: list[str] = []
    for part in parts:
        text = clean(part)
        if text and text not in seen:
            seen.add(text)
            ordered.append(text)
    return sep.join(ordered)


def normalize_continent(region: str, country: str) -> str:
    region_text = clean(region).lower()
    country_text = clean(country).lower()
    if any(key in region_text for key in ["europe", "western europe", "eastern europe"]):
        return "Europa"
    if any(key in region_text for key in ["north america", "caribbean"]):
        return "Ameryka Polnocna"
    if any(key in region_text for key in ["latin america", "south america"]):
        return "Ameryka Poludniowa"
    if any(key in region_text for key in ["africa", "sub-saharan", "north africa"]):
        return "Afryka"
    if any(key in region_text for key in ["oceania", "australia"]):
        return "Australia i Oceania"
    if any(key in region_text for key in ["asia", "middle east"]):
        return "Azja"
    middle_east = {"united arab emirates", "saudi arabia", "qatar", "oman", "kuwait", "bahrain", "israel", "jordan", "lebanon"}
    return "Azja" if country_text in middle_east else ""


def http_get(url: str, timeout: int) -> requests.Response:
    last_response: requests.Response | None = None
    for headers in (REQUEST_HEADERS, None):
        response = requests.get(url, headers=headers, timeout=timeout)
        if response.status_code == 200:
            return response
        last_response = response
        time.sleep(1.0)
    assert last_response is not None
    last_response.raise_for_status()
    return last_response


def fetch_event_sitemap_urls() -> list[str]:
    response = http_get(SITEMAP_INDEX_URL, timeout=45)
    root = ET.fromstring(response.text)
    urls = [node.text.strip() for node in root.findall("sm:sitemap/sm:loc", NS) if node.text]
    return [url for url in urls if re.search(r"/events(?:-\d+)?\.xml$", url)]


def build_public_inventory() -> list[dict[str, str]]:
    seen: set[str] = set()
    rows: list[dict[str, str]] = []
    for sitemap_url in fetch_event_sitemap_urls():
        response = http_get(sitemap_url, timeout=90)
        time.sleep(0.4)
        root = ET.fromstring(response.text)
        sitemap_name = Path(urlparse(sitemap_url).path).name
        for node in root.findall("sm:url/sm:loc", NS):
            url = clean(node.text)
            path_parts = [part for part in urlparse(url).path.split("/") if part]
            if len(path_parts) != 1 or url in seen:
                continue
            seen.add(url)
            rows.append({
                "row_number": str(len(rows) + 1),
                "event_slug": path_parts[0],
                "ten_times_url": url,
                "source_sitemap": sitemap_name,
            })
    rows.sort(key=lambda item: item["ten_times_url"])
    for index, row in enumerate(rows, start=1):
        row["row_number"] = str(index)
    return rows


def resolve_paths(args: argparse.Namespace) -> tuple[Path, Path, Path, Path]:
    workspace = Path(args.workspace)
    output_dir = Path(args.output_dir) if args.output_dir else workspace / "output" / "10times_public_global"
    local_csv = Path(args.local_csv) if args.local_csv else workspace / "output" / "whr_global_tradeshows_june2026_merged_latest.csv"
    local_raw = Path(args.local_raw) if args.local_raw else workspace / "output" / "whr_global_tradeshows_june2026_merged_latest_raw.json"
    return workspace, output_dir, local_csv, local_raw


def load_raw_lookup(local_raw_path: Path) -> dict[str, dict]:
    payload = json.loads(local_raw_path.read_text(encoding="utf-8"))
    lookup: dict[str, dict] = {}
    for item in payload.get("rows", []):
        basic = item.get("basic") or {}
        event_id = clean(basic.get("id"))
        if event_id:
            lookup[event_id] = item
    return lookup


def build_schema_rows(local_csv_path: Path, raw_lookup: dict[str, dict]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with local_csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for source_row in reader:
            event_id = clean(source_row.get("event_id"))
            raw_item = raw_lookup.get(event_id, {})
            basic = raw_item.get("basic") or {}
            location = basic.get("eventLocation") or {}
            organizer = basic.get("organizer") or {}
            region_list = location.get("region") or []
            region = clean(region_list[0]) if region_list else ""
            description = clean(basic.get("description"))
            categories = clean(source_row.get("categories"))
            tags = clean(source_row.get("tags"))
            website = first_nonempty(source_row.get("website"), basic.get("website"))
            ten_times_slug = clean(basic.get("10timesEventPageUrl"))
            visitor_profile = unique_join([
                clean(source_row.get("top_audience_designations")),
                clean(source_row.get("top_audience_countries")),
            ])
            exhibitor_profile = unique_join([tags, categories])
            scope = unique_join([categories, tags, description])
            rows.append({
                "PELNA NAZWA WYDARZENIA": clean(source_row.get("event_name")),
                "DATA OD": iso_to_pl(clean(source_row.get("start_date"))),
                "DATA DO": iso_to_pl(clean(source_row.get("end_date"))),
                "EDYCJE": clean(source_row.get("editions")),
                "MIASTO": clean(source_row.get("city")),
                "KRAJ": clean(source_row.get("country")),
                "KONTYNENT": normalize_continent(region, source_row.get("country", "")),
                "ORGANIZATOR NAZWA SKROCONA": short_organizer_name(first_nonempty(source_row.get("organizer_name"), organizer.get("name"))),
                "ORGANIZATOR PELNA NAZWA": first_nonempty(source_row.get("organizer_name"), organizer.get("name")),
                "WYSTAWCY": clean(source_row.get("estimated_exhibitors")),
                "WYSTAWCY Z POLSKI": "",
                "WYSTAWCY Z KRAJU WYDARZENIA": "",
                "WWW": website or (f"https://10times.com/{ten_times_slug}" if ten_times_slug else ""),
                "ZAKRES BRANZOWY TARGOW": scope,
                "PROFIL WYSTAWCY": exhibitor_profile,
                "PROFIL ODWIEDZAJACEGO": visitor_profile,
            })
    rows.sort(key=lambda item: (item["DATA OD"], item["PELNA NAZWA WYDARZENIA"].lower()))
    return rows


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str], delimiter: str = ";") -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter=delimiter)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, rows: list[dict[str, str]], fieldnames: list[str], title: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{sheet.max_row}"
    for column_index, field in enumerate(fieldnames, start=1):
        width = 28
        if field in {"ten_times_url", "WWW", "ZAKRES BRANZOWY TARGOW", "PROFIL WYSTAWCY", "PROFIL ODWIEDZAJACEGO"}:
            width = 65
        elif field in {"ORGANIZATOR PELNA NAZWA"}:
            width = 45
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    if rows:
        table = Table(displayName=f"T_{title}", ref=f"A1:{get_column_letter(len(fieldnames))}{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)
    workbook.save(path)


def write_report(report_path: Path, inventory_csv: Path, inventory_xlsx: Path, schema_csv: Path | None, schema_xlsx: Path | None, inventory_rows: list[dict[str, str]], schema_rows: list[dict[str, str]]) -> None:
    lines = [
        "10times public export report",
        "============================",
        f"Public base event URLs from sitemap: {len(inventory_rows)}",
        f"Schema rows built from local WHR export: {len(schema_rows)}",
        "",
        "Files:",
        f"- {inventory_csv}",
        f"- {inventory_xlsx}",
    ]
    if schema_csv and schema_xlsx:
        lines.extend([
            f"- {schema_csv}",
            f"- {schema_xlsx}",
        ])
    lines.extend([
        "",
        "Notes:",
        "- Inventory contains only base 10times event URLs discovered via public sitemap.",
        "- Rich schema export is generated only when local WHR-aligned CSV and raw JSON are available.",
        "- Public scraping of event detail pages may be blocked by Cloudflare and is intentionally not bypassed.",
    ])
    report_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    args = parse_args()
    _, output_dir, local_csv_path, local_raw_path = resolve_paths(args)
    output_dir.mkdir(parents=True, exist_ok=True)

    inventory_csv = output_dir / "10times_public_inventory_base_urls.csv"
    inventory_xlsx = output_dir / "10times_public_inventory_base_urls.xlsx"
    schema_csv = output_dir / "10times_schema_from_local_whr.csv"
    schema_xlsx = output_dir / "10times_schema_from_local_whr.xlsx"
    report_path = output_dir / "10times_public_export_report.txt"

    inventory_rows = build_public_inventory()
    inventory_fields = ["row_number", "event_slug", "ten_times_url", "source_sitemap"]
    write_csv(inventory_csv, inventory_rows, inventory_fields)
    write_xlsx(inventory_xlsx, inventory_rows, inventory_fields, "Inventory")

    schema_rows: list[dict[str, str]] = []
    schema_csv_out: Path | None = None
    schema_xlsx_out: Path | None = None
    if local_csv_path.exists() and local_raw_path.exists():
        raw_lookup = load_raw_lookup(local_raw_path)
        schema_rows = build_schema_rows(local_csv_path, raw_lookup)
        write_csv(schema_csv, schema_rows, SCHEMA_COLUMNS)
        write_xlsx(schema_xlsx, schema_rows, SCHEMA_COLUMNS, "Wydarzenia")
        schema_csv_out = schema_csv
        schema_xlsx_out = schema_xlsx

    write_report(report_path, inventory_csv, inventory_xlsx, schema_csv_out, schema_xlsx_out, inventory_rows, schema_rows)
    print(f"OK inventory={len(inventory_rows)} schema={len(schema_rows)}")
    print(inventory_xlsx)
    if schema_xlsx_out:
        print(schema_xlsx_out)
    print(report_path)


if __name__ == "__main__":
    main()
